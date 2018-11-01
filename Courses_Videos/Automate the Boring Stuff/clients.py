import openpyxl
from openpyxl import load_workbook
import os
os.chdir('/home/evolveu/Documents/Excel')

workbook = load_workbook('cSpace_Booking.xlsx')
worksheets = workbook.sheetnames

print(worksheets)

sheet = workbook['Clients']  

fName = []
lName = []
issue = []
Name = []
splitName = []

for i in range (2, 21): # need to change, needs to be dynamic
    Name.append(sheet.cell(row = i, column = 1).value)
    splitName.append(Name[i-2].split())
    fName.append(splitName[i-2][0])
    lName.append(splitName[i-2][1])
    
for k in range (2, 21):
    issue.append(sheet.cell(row = i, column = 6).value)

for i in range(0, 19):
	print(" First Name: " + str(fName[i]) + "\n Last Name: " + str(lName[i]) + "\n Issue: " + str(issue[i]) + "\n")

os.chdir('/home/evolveu/code/play5')
clients = open("clients.html").read().format(fName0=fName[0], lName0=lName[0], issue0=issue[0])
print("hello world")