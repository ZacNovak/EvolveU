import openpyxl
from openpyxl import load_workbook
import os
os.chdir('/home/evolveu/Documents/Excel')

workbook = load_workbook('cSpace_Booking.xlsx')
worksheets = workbook.sheetnames

print(worksheets)

sheet = workbook['Clients']  

newlist = []

print('About to begin for loop')

for i in range (2, 21): # need to change, needs to be dynamic
    # print(sheet.cell(row=i, column = 1).value)
    newlist.append(sheet.cell(row=i, column = 1).value)
    # print(newlist)
    
# print('After for loop')
# print(newlist)
# print('After print newlist')

newlist = [el.split(' ') for el in newlist]

for k in range (2, 21):
    newlist.append(sheet.cell(row=k, column = 6).value)
    
print(newlist)

print('hello world')
