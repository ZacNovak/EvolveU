import openpyxl
from openpyxl import load_workbook
import os
os.chdir('/home/evolveu/Documents/Excel')

workbook = load_workbook('cSpace_Booking.xlsx')
worksheets = workbook.sheetnames

print(worksheets)

sheet = workbook['Clients']  

fName = []
lName = []
issue = []
splitName = []
headingRow = []
nameColumnLetter = None
issuesColumnLetter = None

headingRow = sheet['1']

for cell in headingRow:
	if cell.value == "Name":
		nameColumnLetter = cell.column
	if cell.value == "Issues":
		issuesColumnLetter = cell.column

namecol = sheet[nameColumnLetter]
issuecol = sheet[issuesColumnLetter]

for cell in namecol[1:]:
    splitName = cell.va 
			<td>Issues</td>
	</tr>
	"""
for i in range(0, length):
	html = html + """<tr>"""
	html = html + """<td>""" + fName[i]+ """</td>"""
	html = html + """<td>""" + lName[i] + """</td>"""
	html = html + """<td> """ + str(issue[i]) + """</td>"""
	html = html + """</tr>"""

html = html + """</Table>"""

with open("output_html.html","w") as file:
    file.write(html)
    file.close