
import sys
from openpyxl import load_workbook
# from openpyxl import Workbook

workbook = load_workbook('cSpace_Bookingv1.xlsx')

worksheets = workbook.sheetnames
clients_sheet = workbook['Clients'] 
sheet = workbook['2018-07']

def cell_has_value(ws, row, col):
    if row > ws.max_row:
        return False
    if col > ws.max_column:
        return False
    if not ws.cell(row=row, column=col).value:
        return False
    if not ws.cell(row=row, column=col).value.strip():
        return False

    return True

def find_tab(wb, date_s):
	for sheet in wb:
		if sheet.title == date_s[0:7]:
			return sheet


def create_list_of_names(clients_sheet):
	names = []
	splitnames = []
	firstnames = []
	for row in clients_sheet.iter_rows(min_row=2, max_col=1):
		for cell in row: 
			names.append(cell.value)
	for i in range(len(names)):
		splitnames = names[i].split(" ")
		firstnames.append(splitnames[0])
	return firstnames

def names_in_sheet(sheet, names):
	people = []
	for row in sheet.iter_rows(min_row=2, min_col=3):
			for cell in row:
				if(cell.value is not None):
					if not(cell.value in names):
						if not(cell.value in people):
							people.append(cell.value)		
	return people

names = create_list_of_names(clients_sheet)
people = names_in_sheet(sheet, names)
print(names)
print(people)
