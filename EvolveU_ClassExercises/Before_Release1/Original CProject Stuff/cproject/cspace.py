import openpyxl
from openpyxl import load_workbook
import os
os.chdir('/home/evolveu/code/cproject')

workbook = load_workbook('cSpace_Booking_New.xlsx')

worksheets = workbook.sheetnames

print("What date would you like to book? (please put in format yyyy-mm-dd?")
date = input()

def getRowNumber(date):
	sheetFormat = date[:7]
	sheet = None
	for name in worksheets:
		if name == sheetFormat:
			sheet = name 
	workSheet = workbook[sheet]

	theRowNumber = None
	dateColumn = workSheet['A']
	for cell in dateColumn:
		if str(cell.value)[:7] == sheetFormat:
			theRowNumber = cell.row + int(date[8:10]) - 1 
	return sheet, str(theRowNumber)


def getEmptyRooms(sheetName, rowNumber):
	emptyRooms = []
	for cell in workbook[sheetName][rowNumber]:
		if cell.value is None:
			emptyRooms.append(workbook[sheetName][cell.column + "1"].value)
	return emptyRooms


def printRooms(date):
	sheetName, rowNumber = getRowNumber(date)
	print(getEmptyRooms(sheetName, rowNumber))
	return getEmptyRooms(sheetName, rowNumber)


printRooms(date)