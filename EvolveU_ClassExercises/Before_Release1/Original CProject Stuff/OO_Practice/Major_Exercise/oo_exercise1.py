
import openpyxl
from openpyxl import Workbook, worksheet, load_workbook

workbook = load_workbook('cSpace_Bookingv1.xlsx')

sheetclients = workbook['Clients']
sheetfacilities = workbook['Facilities']
sheetrates = workbook['Rates']
sheetmonth = workbook['2018-11']

class Client:

	def __init__(self, credit):
		self.credit = credit

	def createClientList(sheet):
		clientsDictionary = {}
		for row in sheet.iter_rows(min_row=2):
			fullname = row[0].value
			splitname = fullname.split(" ")
			firstname = splitname[0]
			clientsDictionary[firstname] = Client(0)
		return clientsDictionary

	def assigncredits(clientdict, ratedict, facilitiesdict, monthsheet):
		for row in monthsheet.iter_rows(min_row=2, min_col=2):
			for cell in row:
				if cell.value in clientdict:
					if(monthsheet[cell.column][0].value in facilitiesdict):
						roomtype = facilitiesdict[monthsheet[cell.column][0].value]
						rateadd = ratedict[roomtype]
						clientdict[cell.value].credit = clientdict[cell.value].credit + rateadd
		return clientdict

	def printcreditsowed(clientdict):
		for key in clientdict:
			print(str(key) + ' owes ' + str(clientdict[key].credit) + ' credits ')

def createRatesDictionary(sheet):
	ratesDictionary = {}
	for row in sheet.iter_rows(min_row=2):
		ratesDictionary[row[1].value] = row[2].value
	return ratesDictionary

def createFacilitiesDictionary(sheet):
	facilitiesDictionary = {}
	for row in sheet.iter_rows(min_row=2):
		facilitiesDictionary[row[0].value] = row[1].value
	return facilitiesDictionary



#print(Client.createClientList(sheetclients))

ratedict = createRatesDictionary(sheetrates)
facilitiesdict = createFacilitiesDictionary(sheetfacilities)
clientdict = Client.createClientList(sheetclients)
creditdict = Client.assigncredits(clientdict, ratedict, facilitiesdict,sheetmonth)

Client.printcreditsowed(creditdict)