
import unittest
import oo_exercise1
from openpyxl import Workbook, worksheet


class Testoo_exercise(unittest.TestCase):

	def test_client(self):
		Zac = oo_exercise1.Client(0)
		Lorraine = oo_exercise1.Client(20)

		self.assertEqual(0, Zac.credit)
		self.assertEqual(20, Lorraine.credit)

	def test_createClientList(self):
		wb = self.create_test_wb()
		clientsdictionary = oo_exercise1.Client.createClientList(wb["Clients"])

		self.assertEqual(clientsdictionary['Zac'].credit,0)
		self.assertEqual(clientsdictionary['Julia'].credit,0)
		self.assertEqual(clientsdictionary['Larry'].credit,0)

	def test_createRatesDictionary(self):
		wb = self.create_test_wb()
		self.assertEqual(oo_exercise1.createRatesDictionary(wb["Rates"]),{'Meeting':1, 'Desk':5, 'Gallery':3})

		ratesdictionary = oo_exercise1.createRatesDictionary(wb["Rates"])
		self.assertEqual(ratesdictionary["Meeting"],1)
		self.assertEqual(ratesdictionary["Desk"],5)
		self.assertEqual(ratesdictionary["Gallery"],3)

	def test_createFacilitiesDictionary(self):
		wb = self.create_test_wb()
		self.assertEqual(oo_exercise1.createFacilitiesDictionary(wb["Facilities"]),{'Room a':'Meeting', 'Room b':'Desk', 'Room c':'Gallery'})

		facilitiesdictionary = oo_exercise1.createFacilitiesDictionary(wb["Facilities"])
		self.assertEqual(facilitiesdictionary["Room a"],"Meeting")
		self.assertEqual(facilitiesdictionary["Room b"],"Desk")
		self.assertEqual(facilitiesdictionary["Room c"],"Gallery")

	def test_assigncredits(self):
		wb = self.create_test_wb()
		clientsdictionary = oo_exercise1.Client.createClientList(wb["Clients"])
		ratesdictionary = oo_exercise1.createRatesDictionary(wb["Rates"])
		facilitiesdictionary = oo_exercise1.createFacilitiesDictionary(wb["Facilities"])
		assigncredits = oo_exercise1.Client.assigncredits(clientsdictionary, ratesdictionary, facilitiesdictionary, wb["2018-08"])

		self.assertEqual(clientsdictionary['Zac'].credit,6)
		self.assertEqual(clientsdictionary['Julia'].credit,8)
		self.assertEqual(clientsdictionary['Larry'].credit,7)

	@staticmethod
	def create_test_wb():
		wb = Workbook()
		wb.create_sheet("Clients")
		wb.create_sheet("Facilities")
		wb.create_sheet("Rates")
		wb.create_sheet("2018-07")
		wb.create_sheet("2018-08")
		wb.create_sheet("2018-09") 

		wsclients = wb["Clients"]
		wsclients.cell(row=2, column=1).value = 'Larry'
		wsclients.cell(row=3, column=1).value = 'Zac'
		wsclients.cell(row=4, column=1).value = 'Julia'

		wsrates = wb["Rates"]
		wsrates.cell(row=2, column=2).value = 'Meeting'
		wsrates.cell(row=3, column=2).value = 'Desk'
		wsrates.cell(row=4, column=2).value = 'Gallery'

		wsrates.cell(row=2, column=3).value = 1
		wsrates.cell(row=3, column=3).value = 5
		wsrates.cell(row=4, column=3).value = 3

		wsrates = wb["Facilities"]
		wsrates.cell(row=2, column=1).value = 'Room a'
		wsrates.cell(row=3, column=1).value = 'Room b'
		wsrates.cell(row=4, column=1).value = 'Room c'

		wsrates.cell(row=2, column=2).value = 'Meeting'
		wsrates.cell(row=3, column=2).value = 'Desk'
		wsrates.cell(row=4, column=2).value = 'Gallery'


		wscalendar = wb["2018-08"]
		wscalendar.cell(row=1, column=2).value = 'Room a'
		wscalendar.cell(row=1, column=3).value = 'Room b'
		wscalendar.cell(row=1, column=4).value = 'Room c'

		wscalendar.cell(row=2, column=2).value = 'Larry'
		wscalendar.cell(row=2, column=3).value = 'Zac'
		wscalendar.cell(row=2, column=4).value = None

		wscalendar.cell(row=3, column=2).value = 'Zac'
		wscalendar.cell(row=3, column=3).value = 'Larry'
		wscalendar.cell(row=3, column=4).value = None

		wscalendar.cell(row=4, column=2).value = None
		wscalendar.cell(row=4, column=3).value = 'Julia'
		wscalendar.cell(row=4, column=4).value = 'Julia'

		wscalendar.cell(row=5, column=2).value = 'Larry'

		return wb