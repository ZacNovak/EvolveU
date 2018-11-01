
import sys
from openpyxl import load_workbook
# from openpyxl import Workbook

workbook = load_workbook('cSpace_Bookingv1.xlsx')

#worksheets = workbook.sheetnames
#clients_sheet = workbook['Clients'] 
#sheet = workbook['2018-07']

def cell_has_value(ws, row, col):
    if row > ws.max_row:
        return False
    if col > ws.max_column:
        return False
    if not ws.cell(row=row, column=col).value:
        return False
    if not ws.cell(row=row, column=col).value.strip():
        return False

    return True

def find_tab(wb, date_s):
	for sheet in wb:
		if sheet.title == date_s[0:7]:
			return sheet
	return None


def create_list_of_names(sheet):
	names = []
	for cell in sheet[1][1:]:
		names.append(cell.value)
	return names

def names_in_sheet(sheet, names):
	people = []
	for row in sheet.iter_rows(min_row=2, max_col=4, max_row=5):
			for cell in row:
				if(cell.value is not None):
					if not(cell.value in names):
						if not(cell.value in people):
							people.append(cell.value)		
	return people

