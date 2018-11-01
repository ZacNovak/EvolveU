
import unittest
import major_exercise
from openpyxl import Workbook, worksheet



class TestMajorExercise(unittest.TestCase):

	def test_cell_has_value(self):
		wb = Workbook()
		ws = wb.active

		self.assertFalse(major_exercise.cell_has_value(ws, 1, 1))

		ws.cell(row=1, column=1).value = ''
		self.assertFalse(major_exercise.cell_has_value(ws, 1, 1))

		ws.cell(row=1, column=1).value = 'Room a'
		self.assertTrue(major_exercise.cell_has_value(ws, 1, 1))

		ws.cell(row=1, column=1).value = '   '
		self.assertFalse(major_exercise.cell_has_value(ws, 1, 1))

		self.assertFalse(major_exercise.cell_has_value(ws, 100, 100))
		self.assertFalse(major_exercise.cell_has_value(ws, 1, 100))
		self.assertFalse(major_exercise.cell_has_value(ws, 100, 1))

	def test_find_tab(self):
		wb = self.create_test_wb()
		entered_date = "2018-08"

		ws = major_exercise.find_tab(wb, entered_date)
		self.assertIsInstance(ws, worksheet.Worksheet)
		self.assertEqual(entered_date, ws.title)

		self.assertIsNone(major_exercise.find_tab(wb, '2018-10'))

	def test_create_list_of_names(self):
		wb = self.create_test_wb()
		self.assertEqual(major_exercise.create_list_of_names(wb["Clients"]),['Harry','Zac','Julia'])

	def test_names_in_sheet(self):
		wb = self.create_test_wb()
		self.assertEqual(major_exercise.names_in_sheet(wb["2018-08"],['Harry','Zac','Julia']),['Larry','Jack'])

	@staticmethod
	def create_test_wb():
		wb = Workbook()
		wb.create_sheet("Clients")
		wb.create_sheet("2018-07")
		wb.create_sheet("2018-08")
		wb.create_sheet("2018-09")

		wsclients = wb["Clients"]
		wsclients.cell(column=1, row=2 ).value = 'Harry'
		wsclients.cell(column=1, row=3).value = 'Zac'
		wsclients.cell(column=1, row=4).value = 'Julia'

		wscalendar = wb["2018-08"]
		wscalendar.cell(row=1, column=2).value = 'Room a'
		wscalendar.cell(row=1, column=3).value = 'Room b'
		wscalendar.cell(row=1, column=4).value = 'Room c'

		wscalendar.cell(row=2, column=3).value = 'Harry'
		wscalendar.cell(row=2, column=4).value = 'Harry'
		wscalendar.cell(row=2, column=5).value = None

		wscalendar.cell(row=3, column=3).value = 'Zac'
		wscalendar.cell(row=3, column=4).value = 'Larry'
		wscalendar.cell(row=3, column=5).value = None

		wscalendar.cell(row=4, column=3).value = 'Larry'
		wscalendar.cell(row=4, column=4).value = 'Jack'
		wscalendar.cell(row=4, column=5).value = 'Julia'

		wscalendar.cell(row=5, column=3).value = 'Harry'

        # wb.save('test.xlsx')

		return wb